import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def proc(filename):

	wb = xl.load_workbook(filename)
	sheet = wb['Sheet'] #استبدله باسم الشيت


	for row in range(2,sheet.max_row+1):
    		cell = sheet.cell(row,3)
    		corr_price = cell.value * .9
   		 corr_price_cell = sheet.cell(row,4)
    		corr_price_cell.value = corr_price


	values = Reference(sheet,
         		 min_row=2,
        		  max_row=sheet.max_row,
         		 min_col=4,
         		 max_col=4)
	chart = BarChart()
	chart.add_data(values)
	sheet.add_chart(chart,'e2')
	wb.save(filename)